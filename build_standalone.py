#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成完整优化版周会数据看板 HTML
功能：
- 10 个模块，组合图（柱状+折线双Y轴）
- 棒棒糖图（达成率从低到高排序，颜色分段）
- BC区平均工时 & 超八点下班对比图
- 所有数据表格支持搜索
- 纵坐标动态范围
"""

import json, sys, os, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tdoc_data.json')
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html')

# 优先从 tdoc_data.json（腾讯文档同步数据）读取
try:
    with open(EXCEL_JSON, 'r', encoding='utf-8') as f:
        jd = json.load(f)
    excel_data = {k: v['data'] for k, v in jd['sheets'].items()}
    print(f"Read tdoc_data.json: {len(excel_data)} sheets (source: {jd.get('source', 'unknown')})")
except Exception as e:
    print(f"tdoc_data.json read failed: {e}, falling back to Excel")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), '周会数据模板.xlsx'), data_only=True)
        def safe_float(v):
            if v is None: return None
            try: return float(v)
            except: return str(v).strip() if v else None
        excel_data = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                rows.append([safe_float(c) for c in row])
            excel_data[name] = [r for r in rows if any(c is not None for c in r)]
        print(f"Read Excel: {len(excel_data)} sheets")
    except Exception as e2:
        print(f"Excel read also failed: {e2}")
        sys.exit(1)

# 序列化数据
data_json = json.dumps(excel_data, ensure_ascii=False)

HTML = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>周会数据看板</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
<style>
:root{--bg:#0f172a;--card:#1e293b;--border:#334155;--text:#e2e8f0;--muted:#94a3b8;--accent:#3b82f6;--green:#10b981;--amber:#f59e0b;--red:#ef4444;--purple:#8b5cf6;}
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Microsoft YaHei','PingFang SC',sans-serif;background:var(--bg);color:var(--text);display:flex;min-height:100vh;}
.sidebar{width:220px;background:var(--card);border-right:1px solid var(--border);padding:16px 0;position:fixed;top:0;left:0;bottom:0;overflow-y:auto;z-index:100;}
.sidebar h2{font-size:15px;padding:0 16px;margin-bottom:4px;color:var(--accent);}
.sidebar .subtitle{font-size:11px;color:var(--muted);padding:0 16px;margin-bottom:12px;}
.sidebar .btn{display:block;margin:0 16px 8px;padding:8px 12px;border:none;border-radius:6px;cursor:pointer;font-size:13px;width:calc(100% - 32px);text-align:center;font-weight:600;}
.sidebar .btn-sync{background:var(--accent);color:#fff;}
.sidebar .btn-sync:hover{opacity:0.9;}
.sidebar .btn-sync:disabled{opacity:0.5;cursor:not-allowed;}
.sidebar .status{font-size:11px;color:var(--muted);padding:4px 16px;margin-bottom:8px;line-height:1.5;}
.sidebar nav{display:flex;flex-direction:column;}
.sidebar nav a{color:var(--muted);text-decoration:none;padding:7px 16px;font-size:13px;transition:all .2s;border-left:3px solid transparent;cursor:pointer;}
.sidebar nav a:hover,.sidebar nav a.active{color:var(--text);background:rgba(59,130,246,.1);border-left-color:var(--accent);}
.main{margin-left:220px;flex:1;padding:20px;max-width:calc(100vw - 220px);}
.section{display:none;}
.section.active{display:block;}
.section h3{font-size:20px;margin-bottom:16px;color:var(--text);border-bottom:2px solid var(--accent);padding-bottom:8px;}
.chart-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px;}
.chart-row.full{grid-template-columns:1fr;}
.chart-row.triple{grid-template-columns:1fr 1fr 1fr;}
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px;}
.chart-card h4{font-size:14px;color:var(--muted);margin-bottom:12px;}
.chart-box{width:100%;height:400px;}
.chart-box.tall{height:500px;}
.chart-box.xtall{height:600px;}
.table-wrap{overflow-x:auto;margin-top:8px;max-height:500px;overflow-y:auto;}
table{width:100%;border-collapse:collapse;font-size:12px;}
th,td{padding:5px 8px;text-align:center;border:1px solid var(--border);white-space:nowrap;}
th{background:var(--accent);color:#fff;font-weight:600;position:sticky;top:0;z-index:1;}
td{color:var(--text);}
tr:nth-child(even) td{background:rgba(30,41,59,.5);}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:16px;}
.kpi-card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px;text-align:center;}
.kpi-card .label{font-size:12px;color:var(--muted);margin-bottom:4px;}
.kpi-card .value{font-size:26px;font-weight:700;color:var(--accent);}
.kpi-card .sub{font-size:11px;color:var(--muted);margin-top:4px;}
.search-box{margin-bottom:12px;}
.search-box input{width:100%;padding:8px 12px;background:var(--bg);border:1px solid var(--border);border-radius:6px;color:var(--text);font-size:14px;outline:none;}
.search-box input:focus{border-color:var(--accent);}
@media(max-width:768px){body{flex-direction:column;}.sidebar{width:100%;height:auto;position:relative;padding:12px;}.sidebar nav{flex-direction:row;flex-wrap:wrap;gap:4px;}.sidebar nav a{border-left:none;border-bottom:2px solid transparent;padding:6px 10px;}.sidebar nav a.active{border-bottom-color:var(--accent);}.main{margin-left:0;max-width:100%;padding:12px;}.chart-row,.chart-row.triple{grid-template-columns:1fr;}.chart-box,.chart-box.tall,.chart-box.xtall{height:350px;}}
</style>
</head>
<body>
<div class="sidebar">
  <h2>周会数据看板</h2>
  <div class="subtitle">整车制造部</div>
  <button class="btn btn-sync" onclick="syncFromTencentDoc()" id="btnSync" style="display:none">☁️ 从腾讯文档同步</button>
  <div class="status" id="status">数据已加载</div>
  <nav id="nav"></nav>
</div>
<div class="main" id="main"></div>

<script>
// ====== 内嵌数据 ======
var ALL_DATA = __DATA_PLACEHOLDER__;
var EMBEDDED_DATA = ALL_DATA; // 保留内嵌数据作为备用

// ====== 实时同步 ======
var SERVER_URL = 'http://localhost:8765';
var AUTO_REFRESH_INTERVAL = 2 * 60 * 1000; // 2分钟自动拉取最新数据
var syncInProgress = false;
var serverAvailable = false;

// 显示状态
function setStatus(msg, isGood) {
  var st = document.getElementById('status');
  st.innerHTML = msg;
  st.style.color = isGood === false ? 'var(--red)' : (isGood === true ? 'var(--green)' : 'var(--muted)');
}

// 从服务器加载最新数据
async function loadFromServer() {
  try {
    var resp = await fetch(SERVER_URL + '/api/data');
    if (!resp.ok) throw new Error('HTTP ' + resp.status);
    var jd = await resp.json();
    if (jd && jd.sheets) {
      // 转换格式: sheets[name].data → ALL_DATA[name]
      var newData = {};
      Object.keys(jd.sheets).forEach(function(k) {
        newData[k] = jd.sheets[k].data;
      });
      ALL_DATA = newData;
      serverAvailable = true;
      setStatus('实时数据 ' + (jd.generated_at || ''), true);
      return true;
    }
  } catch(e) {
    serverAvailable = false;
    ALL_DATA = EMBEDDED_DATA;
    setStatus('离线模式（内嵌数据）<br><small>双击「启动服务器.bat」获取实时数据</small>', false);
  }
  return false;
}

// 从腾讯文档同步（调服务器执行 tdoc_sync.py）
async function syncFromTencentDoc() {
  var btn = document.getElementById('btnSync');
  if (syncInProgress) return;
  syncInProgress = true;
  btn.disabled = true;
  btn.textContent = '⏳ 同步中...';
  setStatus('正在从腾讯文档拉取最新数据...');

  try {
    var resp = await fetch(SERVER_URL + '/sync');
    var result = await resp.json();
    if (result.success) {
      setStatus('同步成功！刷新数据中...', true);
      // 同步完成后重新拉取数据并渲染（无需刷新页面）
      await loadFromServer();
      renderAll();
      showSection(curSec);
      btn.textContent = '☁️ 从腾讯文档同步';
    } else {
      var errors = result.steps.filter(function(s) { return !s.success; });
      setStatus('同步失败: ' + (errors.length > 0 ? (errors[0].error || '未知错误') : '未知错误'), false);
      btn.textContent = '☁️ 从腾讯文档同步';
    }
  } catch(e) {
    setStatus('无法连接服务器<br><small>请双击「启动服务器.bat」</small>', false);
    btn.textContent = '☁️ 从腾讯文档同步';
  }
  btn.disabled = false;
  syncInProgress = false;
}

// 自动刷新数据
async function autoRefresh() {
  if (!serverAvailable) {
    // 尝试重新连接服务器
    await loadFromServer();
    if (serverAvailable) {
      renderAll();
      showSection(curSec);
    }
    return;
  }
  var changed = await loadFromServer();
  if (changed) {
    renderAll();
    showSection(curSec);
  }
}

// 页面初始化
(async function init() {
  // 检测运行环境
  var isGitHubPages = window.location.hostname.includes('github.io');
  var isLocal = window.location.hostname === 'localhost' || window.location.hostname === '127.0.0.1';

  if (isGitHubPages) {
    // GitHub Pages: 使用内嵌数据，每15分钟自动同步（由 GitHub Actions 驱动）
    ALL_DATA = EMBEDDED_DATA;
    serverAvailable = false;
    setStatus('数据每15分钟自动同步（GitHub Actions）', true);
    document.getElementById('btnSync').style.display = 'none';
    buildNav();
    renderAll();
    showSection(curSec);
  } else if (isLocal) {
    // 本地环境：尝试连接服务器，显示同步按钮
    document.getElementById('btnSync').style.display = 'inline-block';
    await loadFromServer();
    buildNav();
    renderAll();
    showSection(curSec);
    setInterval(autoRefresh, AUTO_REFRESH_INTERVAL);
  } else {
    // 其他环境（如 localhost.run 隧道）：也尝试连接服务器
    document.getElementById('btnSync').style.display = 'inline-block';
    await loadFromServer();
    buildNav();
    renderAll();
    showSection(curSec);
    setInterval(autoRefresh, AUTO_REFRESH_INTERVAL);
  }
})();

// ====== 工具函数 ======
function fmtNum(n){if(n===null||n===undefined||n==='')return'--';if(typeof n==='string')return n;return n.toLocaleString('zh-CN');}
function getSheet(name){return ALL_DATA[name]||null;}
function num(v){if(v===null||v===undefined||v==='')return NaN;return+v;}
function rateYRange(rates){var valid=rates.filter(function(v){return v!==null&&!isNaN(v)&&v!==0;});if(valid.length===0)return{min:0,max:120};var mn=Math.floor(Math.min.apply(null,valid.concat([70]))/5)*5;var mx=Math.ceil(Math.max.apply(null,valid.concat([100]))/5)*5+5;return{min:mn,max:mx};}

// ====== 导航 ======
var SECTIONS=[
  {id:'s1',name:'月度计划汇总'},
  {id:'s2',name:'周计划汇总'},
  {id:'s3',name:'各线体计划达成率'},
  {id:'s4',name:'BC区工时汇总'},
  {id:'s5',name:'10H产能利用率'},
  {id:'s6',name:'11H产能利用率'},
  {id:'s7',name:'11H排产负荷率'},
  {id:'s8',name:'10H排产负荷率'},
  {id:'s9',name:'切线次数'},
  {id:'s10',name:'25-26年切线对比'}
];
var curSec='s1';

function buildNav(){
  document.getElementById('nav').innerHTML=SECTIONS.map(function(s,i){
    return '<a data-id="'+s.id+'" onclick="showSection(\''+s.id+'\')"'+(i===0?' class="active"':'')+'>'+s.name+'</a>';
  }).join('');
}

function showSection(id){
  curSec=id;
  var secs=document.querySelectorAll('.section');
  for(var i=0;i<secs.length;i++)secs[i].classList.remove('active');
  var links=document.querySelectorAll('nav a');
  for(var i=0;i<links.length;i++)links[i].classList.remove('active');
  var sec=document.getElementById(id);if(sec)sec.classList.add('active');
  var link=document.querySelector('nav a[data-id="'+id+'"]');if(link)link.classList.add('active');
  setTimeout(function(){window.dispatchEvent(new Event('resize'));},150);
}

// ====== 搜索过滤 ======
function filterTable(input, tableId){
  var q=input.value.toLowerCase();
  var rows=document.querySelectorAll('#'+tableId+' tbody tr');
  for(var i=0;i<rows.length;i++){
    var line=rows[i].getAttribute('data-line')||'';
    rows[i].style.display=line.toLowerCase().indexOf(q)>=0?'':'none';
  }
}

// ====== 渲染入口 ======
function renderAll(){
  var html='';
  html+=renderMonthly();
  html+=renderWeekly();
  html+=renderLineRate();
  html+=renderBCLabor();
  html+=renderCapUtil('10H产能利用率','s5','10H');
  html+=renderCapUtil('11H产能利用率','s6','11H');
  html+=renderLoadRate('11H排产负荷率','s7','11H');
  html+=renderLoadRate('10H排产负荷率','s8','10H');
  html+=renderChangeover();
  html+=renderChangeoverYoY();
  document.getElementById('main').innerHTML=html;
  initAllCharts();
}

// ====== 组合图 ======
function makeComboChart(domId,labels,bar1,bar1Name,bar2,bar2Name,lineData,lineName,isRate){
  var dom=document.getElementById(domId);if(!dom)return;
  var chart=echarts.init(dom);
  var yr=isRate?rateYRange(lineData):null;
  var series=[
    {name:bar1Name,type:'bar',data:bar1,itemStyle:{color:'#3b82f6'},barGap:'30%'},
    {name:bar2Name,type:'bar',data:bar2,itemStyle:{color:'#10b981'}},
    {name:lineName,type:'line',yAxisIndex:1,data:lineData,itemStyle:{color:'#f59e0b'},lineStyle:{width:2.5},symbol:'circle',symbolSize:8}
  ];
  var yAxis=[
    {type:'value',name:bar1Name.indexOf('排产')>=0?'排产量':'产出量',nameTextStyle:{color:'#94a3b8'},axisLabel:{color:'#94a3b8',formatter:function(v){return v>=1000?(v/1000).toFixed(0)+'k':v;}}},
    {type:'value',name:'%',nameTextStyle:{color:'#f59e0b'},axisLabel:{color:'#f59e0b',formatter:'{value}%'},min:yr?yr.min:0,max:yr?yr.max:120,splitLine:{show:false}}
  ];
  chart.setOption({
    tooltip:{trigger:'axis'},
    legend:{data:[bar1Name,bar2Name,lineName],textStyle:{color:'#94a3b8'},top:0},
    grid:{left:80,right:80,top:40,bottom:60},
    xAxis:{type:'category',data:labels,axisLabel:{color:'#94a3b8',rotate:labels.length>12?45:0}},
    yAxis:yAxis,series:series
  });
}

// ====== 1. 月度计划汇总 (年生产计划数据汇总) ======
function renderMonthly(){
  var data=getSheet('年生产计划数据汇总');
  if(!data)return'<div class="section" id="s1"><h3>月度计划汇总</h3><p style="color:#ef4444">数据不可用</p></div>';
  var months=data[0].slice(1).filter(function(n){return n;});
  var plans=data[1].slice(1,months.length+1).map(num);
  var actuals=data[2].slice(1,months.length+1).map(num);
  var diffs=data[3].slice(1,months.length+1).map(num);
  var rates=data[4].slice(1,months.length+1).map(function(v){var n=num(v);return isNaN(n)?null:+(n*100).toFixed(2);});
  var tp=plans.reduce(function(a,b){return a+b;},0);
  var ta=actuals.reduce(function(a,b){return a+b;},0);
  var rate=tp>0?(ta/tp*100).toFixed(2):'--';
  var diff=ta-tp;

  var h='<div class="section active" id="s1"><h3>月度计划汇总（1-7月）</h3>';
  h+='<div class="kpi-grid">';
  h+='<div class="kpi-card"><div class="label">计划总量</div><div class="value">'+fmtNum(tp)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">实际产出</div><div class="value">'+fmtNum(ta)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">完成比率</div><div class="value">'+rate+'%</div></div>';
  h+='<div class="kpi-card"><div class="label">差额</div><div class="value" style="color:'+(diff>=0?'var(--green)':'var(--red)')+'">'+(diff>=0?'+':'')+fmtNum(diff)+'</div></div>';
  h+='</div>';
  h+='<div class="chart-row full"><div class="chart-card"><h4>月度计划 vs 实际产出 & 达成率</h4><div class="chart-box tall" id="chart_s1"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="search-box"><input type="text" placeholder="搜索月份..." oninput="filterTable(this,\'tbl_s1\')"></div><div class="table-wrap"><table id="tbl_s1"><thead><tr><th>月份</th><th>计划排产</th><th>实际产出</th><th>差额</th><th>计划达成率</th></tr></thead><tbody>';
  for(var i=0;i<months.length;i++){
    h+='<tr data-line="'+months[i]+'"><td>'+months[i]+'</td><td>'+fmtNum(plans[i])+'</td><td>'+fmtNum(actuals[i])+'</td><td style="color:'+(diffs[i]>=0?'var(--green)':'var(--red)')+'">'+(diffs[i]>=0?'+':'')+fmtNum(diffs[i])+'</td><td>'+(rates[i]!==null?rates[i].toFixed(2)+'%':'--')+'</td></tr>';
  }
  h+='<tr style="font-weight:700"><td>合计</td><td>'+fmtNum(tp)+'</td><td>'+fmtNum(ta)+'</td><td style="color:'+(diff>=0?'var(--green)':'var(--red)')+'">'+(diff>=0?'+':'')+fmtNum(diff)+'</td><td>'+rate+'%</td></tr>';
  h+='</tbody></table></div></div></div>';
  window._s1={labels:months,plans:plans,actuals:actuals,diffs:diffs,rates:rates};
  return h;
}

// ====== 2. 周计划汇总 (月生产计划数据汇总) ======
function renderWeekly(){
  var data=getSheet('月生产计划数据汇总');
  if(!data)return'<div class="section" id="s2"><h3>周计划汇总</h3><p style="color:#ef4444">数据不可用</p></div>';
  var rawWeeks=data[0].slice(1);
  var rawPlans=data[1].slice(1).map(num);
  var rawActuals=data[2].slice(1).map(num);
  var rawDiffs=data[3].slice(1).map(num);
  var rawRates=data[4].slice(1).map(function(v){var n=num(v);return isNaN(n)?null:+(n*100).toFixed(2);});
  var weeks=[],plans=[],actuals=[],diffs=[],rates=[];
  for(var i=0;i<rawWeeks.length;i++){
    if(!isNaN(rawPlans[i])&&!isNaN(rawActuals[i])){
      weeks.push(rawWeeks[i]);plans.push(rawPlans[i]);actuals.push(rawActuals[i]);
      diffs.push(rawDiffs[i]);rates.push(rawRates[i]);
    }
  }
  var tp=plans.reduce(function(a,b){return a+b;},0);
  var ta=actuals.reduce(function(a,b){return a+b;},0);
  var rate=tp>0?(ta/tp*100).toFixed(2):'--';
  var diff=ta-tp;

  var h='<div class="section" id="s2"><h3>周计划汇总</h3>';
  h+='<div class="kpi-grid">';
  h+='<div class="kpi-card"><div class="label">计划总量</div><div class="value">'+fmtNum(tp)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">实际产出</div><div class="value">'+fmtNum(ta)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">完成比率</div><div class="value">'+rate+'%</div></div>';
  h+='<div class="kpi-card"><div class="label">差额</div><div class="value" style="color:'+(diff>=0?'var(--green)':'var(--red)')+'">'+(diff>=0?'+':'')+fmtNum(diff)+'</div></div>';
  h+='</div>';
  h+='<div class="chart-row full"><div class="chart-card"><h4>周计划 vs 实际产出 & 达成率</h4><div class="chart-box tall" id="chart_s2"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="table-wrap"><table><thead><tr><th>周</th><th>计划排产</th><th>实际产出</th><th>差额</th><th>计划达成率</th></tr></thead><tbody>';
  for(var i=0;i<weeks.length;i++){
    h+='<tr><td>'+weeks[i]+'</td><td>'+fmtNum(plans[i])+'</td><td>'+fmtNum(actuals[i])+'</td><td style="color:'+(diffs[i]>=0?'var(--green)':'var(--red)')+'">'+(diffs[i]>=0?'+':'')+fmtNum(diffs[i])+'</td><td>'+(rates[i]!==null?rates[i].toFixed(2)+'%':'--')+'</td></tr>';
  }
  h+='<tr style="font-weight:700"><td>合计</td><td>'+fmtNum(tp)+'</td><td>'+fmtNum(ta)+'</td><td style="color:'+(diff>=0?'var(--green)':'var(--red)')+'">'+(diff>=0?'+':'')+fmtNum(diff)+'</td><td>'+rate+'%</td></tr>';
  h+='</tbody></table></div></div></div>';
  window._s2={labels:weeks,plans:plans,actuals:actuals,diffs:diffs,rates:rates};
  return h;
}

// ====== 3. 各线体计划达成率 ======
function renderLineRate(){
  var data=getSheet('各线体周生产数据汇总');
  if(!data)return'<div class="section" id="s3"><h3>各线体计划达成率</h3><p style="color:#ef4444">数据不可用</p></div>';
  var lineNames=data[0].slice(1).filter(function(n){return n;});
  var plans=data[1].slice(1,lineNames.length+1).map(num);
  var actuals=data[2].slice(1,lineNames.length+1).map(num);
  var preRates=data[4]?data[4].slice(1,lineNames.length+1).map(function(v){var n=num(v);return isNaN(n)?null:+(n*100).toFixed(2);}):[];
  var tp=plans.reduce(function(a,b){return a+b;},0);
  var ta=actuals.reduce(function(a,b){return a+b;},0);
  var rate=tp>0?(ta/tp*100).toFixed(2):'--';
  var diff=ta-tp;

  var h='<div class="section" id="s3"><h3>各线体计划达成率</h3>';
  h+='<div class="kpi-grid">';
  h+='<div class="kpi-card"><div class="label">计划总量</div><div class="value">'+fmtNum(tp)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">实际产出</div><div class="value">'+fmtNum(ta)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">达成率</div><div class="value">'+rate+'%</div></div>';
  h+='<div class="kpi-card"><div class="label">差额</div><div class="value" style="color:'+(diff>=0?'var(--green)':'var(--red)')+'">'+(diff>=0?'+':'')+fmtNum(diff)+'</div></div>';
  h+='</div>';
  h+='<div class="chart-row"><div class="chart-card"><h4>计划 vs 生产 & 达成率（组合图）</h4><div class="chart-box tall" id="chart_s3a"></div></div>';
  h+='<div class="chart-card"><h4>达成率排序（棒棒糖图）</h4><div class="chart-box tall" id="chart_s3b"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="search-box"><input type="text" placeholder="搜索线体..." oninput="filterTable(this,\'tbl_s3\')"></div><div class="table-wrap"><table id="tbl_s3"><thead><tr><th>线体</th><th>计划总量</th><th>生产总量</th><th>差额</th><th>计划达成率</th></tr></thead><tbody>';
  for(var i=0;i<lineNames.length;i++){
    var d=actuals[i]-plans[i];
    h+='<tr data-line="'+lineNames[i]+'"><td>'+lineNames[i]+'</td><td>'+fmtNum(plans[i])+'</td><td>'+fmtNum(actuals[i])+'</td><td style="color:'+(d>=0?'var(--green)':'var(--red)')+'">'+(d>=0?'+':'')+fmtNum(d)+'</td><td>'+(preRates[i]!==null?preRates[i].toFixed(2)+'%':'--')+'</td></tr>';
  }
  h+='<tr style="font-weight:700"><td>合计</td><td>'+fmtNum(tp)+'</td><td>'+fmtNum(ta)+'</td><td style="color:'+(diff>=0?'var(--green)':'var(--red)')+'">'+(diff>=0?'+':'')+fmtNum(diff)+'</td><td>'+rate+'%</td></tr>';
  h+='</tbody></table></div></div></div>';
  window._s3={labels:lineNames,plans:plans,actuals:actuals,preRates:preRates};
  return h;
}

// ====== 4. BC区工时汇总 ======
function renderBCLabor(){
  var data=getSheet('BC区工时汇总');
  if(!data)return'<div class="section" id="s4"><h3>BC区工时汇总</h3><p style="color:#ef4444">数据不可用</p></div>';
  var bLines=[],bHours=[],bDays=[],bOvertime=[],bAvg=[];
  var cLines=[],cHours=[],cDays=[],cOvertime=[],cAvg=[];
  for(var i=2;i<data.length;i++){
    var row=data[i];
    var col0=String(row[0]||'').trim();
    var col6=String(row[6]||'').trim();
    if(col0&&col0!=='合计'&&col0!=='总计'){
      bLines.push(col0);bHours.push(num(row[1])||0);bDays.push(num(row[2])||0);bOvertime.push(num(row[3])||0);bAvg.push(num(row[4])||0);
    }
    if(col6&&col6!=='合计'&&col6!=='总计'){
      cLines.push(col6);cHours.push(num(row[7])||0);cDays.push(num(row[8])||0);cOvertime.push(num(row[9])||0);cAvg.push(num(row[10])||0);
    }
  }
  // 匹配 B区和C区共有的线体
  var tblBC=function(title,lines,hours,days,overtime,avg,tableId){
    var h='<div class="chart-card"><h4>'+title+'</h4><div class="search-box"><input type="text" placeholder="搜索线体..." oninput="filterTable(this,\''+tableId+'\')"></div><div class="table-wrap"><table id="'+tableId+'"><thead><tr><th>线体</th><th>总工时</th><th>出勤天数</th><th>超8点下班</th><th>平均时长</th></tr></thead><tbody>';
    var th=hours.reduce(function(a,b){return a+b;},0);
    var td=days.reduce(function(a,b){return a+b;},0);
    var to2=overtime.reduce(function(a,b){return a+b;},0);
    for(var i=0;i<lines.length;i++){
      h+='<tr data-line="'+lines[i]+'"><td>'+lines[i]+'</td><td>'+hours[i]+'</td><td>'+days[i]+'</td><td>'+overtime[i]+'</td><td>'+avg[i]+'</td></tr>';
    }
    h+='<tr style="font-weight:700"><td>合计</td><td>'+th+'</td><td>'+td+'</td><td>'+to2+'</td><td>'+(td>0?(th/td).toFixed(1):'--')+'</td></tr>';
    h+='</tbody></table></div></div>';
    return h;
  };

  var h='<div class="section" id="s4"><h3>BC区工时汇总</h3>';
  h+='<div class="chart-row"><div class="chart-card"><h4>B区总工时</h4><div class="chart-box tall" id="chart_s4a"></div></div>';
  h+='<div class="chart-card"><h4>C区总工时</h4><div class="chart-box tall" id="chart_s4b"></div></div></div>';
  h+=tblBC('B区工时明细',bLines,bHours,bDays,bOvertime,bAvg,'tbl_s4b');
  h+=tblBC('C区工时明细',cLines,cHours,cDays,cOvertime,cAvg,'tbl_s4c');
  h+='</div>';
  window._s4={bLines:bLines,bHours:bHours,bDays:bDays,bOvertime:bOvertime,bAvg:bAvg,
    cLines:cLines,cHours:cHours,cDays:cDays,cOvertime:cOvertime,cAvg:cAvg};
  return h;
}

// ====== 5/6. 产能利用率 ======
function renderCapUtil(sheetName,secId,label){
  var data=getSheet(sheetName);
  if(!data)return'<div class="section" id="'+secId+'"><h3>'+label+'产能利用率</h3><p style="color:#ef4444">数据不可用</p></div>';
  var lineNames=data[0].slice(1).filter(function(n){return n;});
  var theory=data[1].slice(1,lineNames.length+1).map(num);
  var actual=data[2].slice(1,lineNames.length+1).map(num);
  var utilRates=data[3]?data[3].slice(1,lineNames.length+1).map(function(v){var n=num(v);return isNaN(n)?null:+(n*100).toFixed(2);}):[];
  var tt=theory.reduce(function(a,b){return a+b;},0);
  var ta=actual.reduce(function(a,b){return a+b;},0);
  var h='<div class="section" id="'+secId+'"><h3>'+label+'产能利用率</h3>';
  h+='<div class="kpi-grid">';
  h+='<div class="kpi-card"><div class="label">理论产出</div><div class="value">'+fmtNum(tt)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">实际产出</div><div class="value">'+fmtNum(ta)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">利用率</div><div class="value">'+(tt>0?(ta/tt*100).toFixed(2):'--')+'%</div></div>';
  h+='</div>';
  h+='<div class="chart-row full"><div class="chart-card"><h4>'+label+' 理论/实际产出 & 利用率</h4><div class="chart-box tall" id="chart_'+secId+'"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="search-box"><input type="text" placeholder="搜索线体..." oninput="filterTable(this,\'tbl_'+secId+'\')"></div><div class="table-wrap"><table id="tbl_'+secId+'"><thead><tr><th>线体</th><th>理论产出</th><th>实际产出</th><th>产能利用率</th></tr></thead><tbody>';
  for(var i=0;i<lineNames.length;i++){
    h+='<tr data-line="'+lineNames[i]+'"><td>'+lineNames[i]+'</td><td>'+fmtNum(theory[i])+'</td><td>'+fmtNum(actual[i])+'</td><td>'+(utilRates[i]!==null?utilRates[i].toFixed(2)+'%':'--')+'</td></tr>';
  }
  h+='<tr style="font-weight:700"><td>合计</td><td>'+fmtNum(tt)+'</td><td>'+fmtNum(ta)+'</td><td>'+(tt>0?(ta/tt*100).toFixed(2):'--')+'%</td></tr>';
  h+='</tbody></table></div></div></div>';
  window['_'+secId]={labels:lineNames,theory:theory,actual:actual,utilRates:utilRates};
  return h;
}

// ====== 7/8. 排产负荷率 ======
function renderLoadRate(sheetName,secId,label){
  var data=getSheet(sheetName);
  if(!data)return'<div class="section" id="'+secId+'"><h3>'+label+'排产负荷率</h3><p style="color:#ef4444">数据不可用</p></div>';
  var lineNames=data[0].slice(1).filter(function(n){return n;});
  var theory=data[1].slice(1,lineNames.length+1).map(num);
  var actual=data[2].slice(1,lineNames.length+1).map(num);
  var loadRates=data[3]?data[3].slice(1,lineNames.length+1).map(function(v){var n=num(v);return isNaN(n)?null:+(n*100).toFixed(2);}):[];
  var redundancy=data[4]?data[4].slice(1,lineNames.length+1).map(function(v){var n=num(v);return isNaN(n)?null:+(n*100).toFixed(2);}):[];
  var tt=theory.reduce(function(a,b){return a+b;},0);
  var ta=actual.reduce(function(a,b){return a+b;},0);
  var h='<div class="section" id="'+secId+'"><h3>'+label+'排产负荷率</h3>';
  h+='<div class="kpi-grid">';
  h+='<div class="kpi-card"><div class="label">理论排产</div><div class="value">'+fmtNum(tt)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">实际排产</div><div class="value">'+fmtNum(ta)+'</div></div>';
  h+='<div class="kpi-card"><div class="label">利用率</div><div class="value">'+(tt>0?(ta/tt*100).toFixed(2):'--')+'%</div></div>';
  h+='</div>';
  h+='<div class="chart-row full"><div class="chart-card"><h4>'+label+' 理论/实际排产 & 负荷率+冗余率</h4><div class="chart-box tall" id="chart_'+secId+'"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="search-box"><input type="text" placeholder="搜索线体..." oninput="filterTable(this,\'tbl_'+secId+'\')"></div><div class="table-wrap"><table id="tbl_'+secId+'"><thead><tr><th>线体</th><th>理论排产</th><th>实际排产</th><th>排产负荷率</th><th>排产冗余率</th></tr></thead><tbody>';
  for(var i=0;i<lineNames.length;i++){
    h+='<tr data-line="'+lineNames[i]+'"><td>'+lineNames[i]+'</td><td>'+fmtNum(theory[i])+'</td><td>'+fmtNum(actual[i])+'</td><td>'+(loadRates[i]!==null?loadRates[i].toFixed(2)+'%':'--')+'</td><td>'+(redundancy[i]!==null?redundancy[i].toFixed(2)+'%':'--')+'</td></tr>';
  }
  h+='</tbody></table></div></div></div>';
  window['_'+secId]={labels:lineNames,theory:theory,actual:actual,loadRates:loadRates,redundancy:redundancy};
  return h;
}

// ====== 9. 切线次数 ======
function renderChangeover(){
  var data=getSheet('切线次数');
  if(!data)return'<div class="section" id="s9"><h3>切线次数</h3><p style="color:#ef4444">数据不可用</p></div>';
  var allNames=data[0].slice(1).filter(function(n){return n;});
  var allVals=data[1].slice(1,allNames.length+1).map(num);
  // 分离汇总和普通线体
  var lineNames=[],values=[],totalVal=0;
  for(var i=0;i<allNames.length;i++){
    if(allNames[i]==='汇总'){totalVal=allVals[i]||0;}
    else{lineNames.push(allNames[i]);values.push(allVals[i]||0);}
  }
  var h='<div class="section" id="s9"><h3>切线次数</h3>';
  h+='<div class="kpi-grid"><div class="kpi-card"><div class="label">总切线次数</div><div class="value">'+totalVal+'</div></div></div>';
  h+='<div class="chart-row full"><div class="chart-card"><h4>各线体切线次数</h4><div class="chart-box" id="chart_s9"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="search-box"><input type="text" placeholder="搜索线体..." oninput="filterTable(this,\'tbl_s9\')"></div><div class="table-wrap"><table id="tbl_s9"><thead><tr><th>线体</th><th>切线次数</th></tr></thead><tbody>';
  for(var i=0;i<lineNames.length;i++){
    h+='<tr data-line="'+lineNames[i]+'"><td>'+lineNames[i]+'</td><td>'+(values[i]||0)+'</td></tr>';
  }
  h+='</tbody></table></div></div></div>';
  window._s9={labels:lineNames,values:values};
  return h;
}

// ====== 10. 25-26年切线次数对比 ======
function renderChangeoverYoY(){
  var data=getSheet('25-26年切线次数对比');
  if(!data)return'<div class="section" id="s10"><h3>25-26年切线次数对比</h3><p style="color:#ef4444">数据不可用</p></div>';
  var months=data[0].slice(1).filter(function(n){return n;});
  var y25=data[1].slice(1,months.length+1).map(num);
  var y26=data[2].slice(1,months.length+1).map(num);
  var h='<div class="section" id="s10"><h3>25-26年切线次数对比</h3>';
  h+='<div class="chart-row full"><div class="chart-card"><h4>25-26年切线次数对比</h4><div class="chart-box" id="chart_s10"></div></div></div>';
  h+='<div class="chart-card"><h4>数据明细</h4><div class="search-box"><input type="text" placeholder="搜索月份..." oninput="filterTable(this,\'tbl_s10\')"></div><div class="table-wrap"><table id="tbl_s10"><thead><tr><th>月份</th><th>2025年</th><th>2026年</th></tr></thead><tbody>';
  for(var i=0;i<months.length;i++){
    h+='<tr data-line="'+months[i]+'"><td>'+months[i]+'</td><td>'+fmtNum(y25[i])+'</td><td>'+fmtNum(y26[i])+'</td></tr>';
  }
  h+='</tbody></table></div></div></div>';
  window._s10={labels:months,y25:y25,y26:y26};
  return h;
}

// ====== 图表初始化 ======
function initAllCharts(){
  // 销毁旧图表实例（防止重复渲染时内存泄漏）
  var oldCharts = document.querySelectorAll('.chart-box');
  for (var i = 0; i < oldCharts.length; i++) {
    var instance = echarts.getInstanceByDom(oldCharts[i]);
    if (instance) instance.dispose();
  }
  // 月度汇总
  if(window._s1)makeComboChart('chart_s1',window._s1.labels,window._s1.plans,'计划排产',window._s1.actuals,'实际产出',window._s1.rates,'达成率',true);
  // 周汇总
  if(window._s2)makeComboChart('chart_s2',window._s2.labels,window._s2.plans,'计划排产',window._s2.actuals,'实际产出',window._s2.rates,'达成率',true);
  // 线体达成率
  if(window._s3){
    makeComboChart('chart_s3a',window._s3.labels,window._s3.plans,'计划',window._s3.actuals,'生产',window._s3.preRates,'达成率',true);
    makeLollipopChart('chart_s3b',window._s3.labels,window._s3.preRates);
  }
  // BC区
  if(window._s4){
    makeBarChart('chart_s4a',window._s4.bLines,window._s4.bHours,'B区总工时','#3b82f6');
    makeBarChart('chart_s4b',window._s4.cLines,window._s4.cHours,'C区总工时','#10b981');
  }
  // 产能利用率
  if(window._s5)makeComboChart('chart_s5',window._s5.labels,window._s5.theory,'理论产出',window._s5.actual,'实际产出',window._s5.utilRates,'利用率',true);
  if(window._s6)makeComboChart('chart_s6',window._s6.labels,window._s6.theory,'理论产出',window._s6.actual,'实际产出',window._s6.utilRates,'利用率',true);
  // 排产负荷率
  if(window._s7){
    makeLoadComboChart('chart_s7',window._s7.labels,window._s7.theory,'理论排产',window._s7.actual,'实际排产',window._s7.loadRates,'负荷率',window._s7.redundancy,'冗余率');
  }
  if(window._s8){
    makeLoadComboChart('chart_s8',window._s8.labels,window._s8.theory,'理论排产',window._s8.actual,'实际排产',window._s8.loadRates,'负荷率',window._s8.redundancy,'冗余率');
  }
  // 切线
  if(window._s9)makeBarChart('chart_s9',window._s9.labels,window._s9.values,'切线次数','#8b5cf6');
  // 25-26对比
  if(window._s10)makeDualBarChart('chart_s10',window._s10.labels,window._s10.y25,'2025年',window._s10.y26,'2026年','切线次数');
}

function makeBarChart(domId,labels,values,name,color){
  var dom=document.getElementById(domId);if(!dom)return;
  var chart=echarts.init(dom);
  chart.setOption({
    tooltip:{trigger:'axis'},
    grid:{left:70,right:20,bottom:60},
    xAxis:{type:'category',data:labels,axisLabel:{color:'#94a3b8',rotate:labels.length>12?45:0}},
    yAxis:{type:'value',name:name,nameTextStyle:{color:'#94a3b8'},axisLabel:{color:'#94a3b8'}},
    series:[{type:'bar',data:values,itemStyle:{color:color}}]
  });
}

function makeDualBarChart(domId,labels,values1,name1,values2,name2,ylabel){
  var dom=document.getElementById(domId);if(!dom)return;
  var chart=echarts.init(dom);
  chart.setOption({
    tooltip:{trigger:'axis'},
    legend:{data:[name1,name2],textStyle:{color:'#94a3b8'},top:0},
    grid:{left:70,right:20,top:40,bottom:60},
    xAxis:{type:'category',data:labels,axisLabel:{color:'#94a3b8',rotate:45}},
    yAxis:{type:'value',name:ylabel,nameTextStyle:{color:'#94a3b8'},axisLabel:{color:'#94a3b8'}},
    series:[
      {name:name1,type:'bar',data:values1,itemStyle:{color:'#3b82f6'}},
      {name:name2,type:'bar',data:values2,itemStyle:{color:'#10b981'}}
    ]
  });
}

function makeLoadComboChart(domId,labels,bar1,bar1Name,bar2,bar2Name,lineData,lineName,lineData2,lineName2){
  var dom=document.getElementById(domId);if(!dom)return;
  var chart=echarts.init(dom);
  var yr=rateYRange(lineData.concat(lineData2||[]));
  var series=[
    {name:bar1Name,type:'bar',data:bar1,itemStyle:{color:'#3b82f6'},barGap:'30%'},
    {name:bar2Name,type:'bar',data:bar2,itemStyle:{color:'#10b981'}},
    {name:lineName,type:'line',yAxisIndex:1,data:lineData,itemStyle:{color:'#f59e0b'},lineStyle:{width:2.5},symbol:'circle',symbolSize:7}
  ];
  if(lineData2){
    series.push({name:lineName2,type:'line',yAxisIndex:1,data:lineData2,itemStyle:{color:'#ef4444'},lineStyle:{type:'dotted',width:2},symbol:'diamond',symbolSize:6});
  }
  chart.setOption({
    tooltip:{trigger:'axis'},
    legend:{data:series.map(function(s){return s.name;}),textStyle:{color:'#94a3b8'},top:0},
    grid:{left:80,right:80,top:40,bottom:60},
    xAxis:{type:'category',data:labels,axisLabel:{color:'#94a3b8',rotate:labels.length>12?45:0}},
    yAxis:[
      {type:'value',name:'排产量',axisLabel:{color:'#94a3b8',formatter:function(v){return v>=1000?(v/1000).toFixed(0)+'k':v;}}},
      {type:'value',name:'%',axisLabel:{color:'#f59e0b',formatter:'{value}%'},min:yr.min,max:yr.max,splitLine:{show:false}}
    ],
    series:series
  });
}

// ====== 棒棒糖图 (达成率排序) ======
function makeLollipopChart(domId,labels,values){
  var dom=document.getElementById(domId);if(!dom)return;
  var chart=echarts.init(dom);
  // 过滤有效数据并排序
  var pairs=[];
  for(var i=0;i<labels.length;i++){
    if(values[i]!==null&&!isNaN(values[i])&&values[i]!==0){
      pairs.push({name:labels[i],value:values[i]});
    }
  }
  pairs.sort(function(a,b){return a.value-b.value;});
  var sortedNames=pairs.map(function(p){return p.name;});
  var sortedVals=pairs.map(function(p){return p.value;});

  // 颜色分段
  var colors=sortedVals.map(function(v){
    if(v>=100)return'#10b981';
    if(v>=95)return'#3b82f6';
    if(v>=90)return'#f59e0b';
    return'#ef4444';
  });

  chart.setOption({
    tooltip:{trigger:'axis',formatter:function(p){return p[0].name+'<br/>达成率: '+p[0].value+'%';}},
    grid:{left:70,right:60,top:10,bottom:60},
    xAxis:{type:'value',name:'%',nameTextStyle:{color:'#94a3b8'},axisLabel:{color:'#94a3b8',formatter:'{value}%'},min:Math.floor(Math.min.apply(null,sortedVals.concat([70]))/5)*5,max:Math.ceil(Math.max.apply(null,sortedVals.concat([100]))/5)*5+5},
    yAxis:{type:'category',data:sortedNames,axisLabel:{color:'#94a3b8'},inverse:true},
    series:[
      {type:'bar',data:sortedVals.map(function(v,i){return{value:v,itemStyle:{color:colors[i]}};}),barWidth:12,barGap:'50%'},
      {type:'scatter',data:sortedVals,itemStyle:{color:'#fff',borderColor:'#1e293b',borderWidth:2},symbolSize:14}
    ]
  });
}

// 响应式
window.addEventListener('resize',function(){
  var boxes=document.querySelectorAll('.chart-box');
  for(var i=0;i<boxes.length;i++){var c=echarts.getInstanceByDom(boxes[i]);if(c)c.resize();}
});

// 键盘导航
document.addEventListener('keydown',function(e){
  var n=parseInt(e.key);
  if(n>=0&&n<=9&&SECTIONS[n-1])showSection(SECTIONS[n-1].id);
});

// 启动
(function(){
  buildNav();
  renderAll();
  document.getElementById('status').textContent='数据已加载 | '+Object.keys(ALL_DATA).length+' 个子表';
})();
</script>
</body>
</html>'''

# 替换数据占位符并写入文件
final_html = HTML.replace('__DATA_PLACEHOLDER__', data_json)
with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(final_html)

size_kb = os.path.getsize(OUTPUT) / 1024
print(f"[DONE] {OUTPUT}")
print(f"[SIZE] {size_kb:.1f} KB")
print(f"[SECTIONS] 10 modules with combo charts, lollipop chart, search boxes")
print(f"[INFO] Truly standalone - open directly in any browser!")