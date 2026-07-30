#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
腾讯文档 → 结构化 JSON 同步脚本
1. 读取 Excel 模板了解每个子表的行列结构
2. 从腾讯文档 API 获取最新数据（protobuf 解析）
3. 将原始数据按 Excel 模板结构重组为二维数组
4. 输出 tdoc_data.json 供 HTML 看板加载
"""

import json
import base64
import zlib
import struct
import urllib.request
import sys
import os
import io
import time
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ========== 配置 ==========
DOC_ID = "DU1hxbXJkZFBXRnVw"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, '周会数据模板.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, 'tdoc_data.json')

# 子表 tab ID 映射（从 workbook protobuf 中解析）
TAB_IDS = {
    "年生产计划数据汇总": "BB08J2",
    "月生产计划数据汇总": "gsyuwk",
    "各线体周生产数据汇总": "ztseql",
    "BC区工时汇总": "o5kon0",
    "10H产能利用率": "29zlxb",
    "11H产能利用率": "9s82es",
    "11H排产负荷率": "bulius",
    "10H排产负荷率": "krho19",
    "计划达成率": "vpvsiv",
    "切线次数": "2bu5ok",
    "25-26年切线次数对比": "a65cg9",
}


def safe_float(v):
    if v is None: return None
    try: return float(v)
    except: return str(v).strip() if v else None


def read_excel_template():
    """读取 Excel 模板，获取每个子表的行列结构"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        templates = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                rows.append([safe_float(c) for c in row])
            # 过滤全空行
            rows = [r for r in rows if any(c is not None for c in r)]
            templates[name] = {
                'rows': len(rows),
                'cols': len(rows[0]) if rows else 0,
                'data': rows
            }
        print(f"  Excel template: {len(templates)} sheets loaded")
        return templates
    except Exception as e:
        print(f"  [WARN] Cannot read Excel template: {e}")
        return None


def decode_bytes(val):
    """解码 blackboxprotobuf 的 bytes 字符串"""
    if isinstance(val, str) and val.startswith("b'") and val.endswith("'"):
        try:
            return eval(val).decode('utf-8')
        except:
            return val
    return val


def decode_int64_as_double(val):
    """将 protobuf int64 解码为 IEEE 754 double"""
    if isinstance(val, int) and val > 1000000:
        try:
            return round(struct.unpack('<d', struct.pack('<Q', val))[0], 6)
        except:
            return float(val)
    return float(val) if isinstance(val, (int, float)) else val


def fetch_raw_data(doc_id, tab_id):
    """从腾讯文档 API 获取原始 strings 和 numbers"""
    url = f"https://docs.qq.com/dop-api/opendoc?id={doc_id}&tab={tab_id}&outformat=1&normal=1"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": f"https://docs.qq.com/sheet/{doc_id}"
    }

    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except Exception as e:
        return None, f"Request failed: {e}"

    try:
        cv = data['clientVars']['collab_client_vars']
        iat = cv['initialAttributedText']
        texts = iat['text']
        if not texts:
            return None, "No text entries"
        text = texts[0]
        if 'related_sheet' not in text or not text['related_sheet']:
            return None, "No related_sheet data"

        raw = base64.b64decode(text['related_sheet'])
        decompressed = zlib.decompress(raw)

        import blackboxprotobuf
        result, _ = blackboxprotobuf.decode_message(decompressed)

        sections = result.get('1', {}).get('5', [])
        f19 = None
        for section in sections:
            if '19' in section:
                f19 = section['19']
                break
        if f19 is None:
            return None, "No field 19 found"

        f5 = f19.get('5', {})

        # Extract strings
        strings_raw = f5.get('1', [])
        strings = []
        for s in strings_raw:
            if isinstance(s, dict) and '1' in s:
                strings.append(decode_bytes(s['1']))
            else:
                strings.append(str(s))

        # Extract numbers
        nums_raw = f5.get('3', [])
        numbers = []
        if isinstance(nums_raw, dict):
            # Single number stored as dict (e.g., {'1': 4639763147354275840})
            for v in nums_raw.values():
                numbers.append(decode_int64_as_double(v))
        elif isinstance(nums_raw, list):
            for n in nums_raw:
                if isinstance(n, dict) and '1' in n:
                    numbers.append(decode_int64_as_double(n['1']))
                else:
                    numbers.append(decode_int64_as_double(n))

        # Extract field 6 cell data (sparse representation for sheets with few rows)
        f6 = f19.get('6', [])
        cell_data = {}
        for entry in f6:
            if not isinstance(entry, dict):
                continue
            row = entry.get('1', None)
            col = entry.get('2', None)
            cell_info = entry.get('3', {})
            if row is not None and col is not None and isinstance(cell_info, dict):
                # Value is in cell_info.2.1 (varint)
                f2 = cell_info.get('2', None)
                val = None
                if isinstance(f2, dict):
                    val = f2.get('1', None)
                elif isinstance(f2, int):
                    val = f2
                if val is not None:
                    cell_data[(row, col)] = val

        return {'strings': strings, 'numbers': numbers, 'cell_data': cell_data}, None

    except Exception as e:
        import traceback
        return None, f"Parse error: {e}\n{traceback.format_exc()}"


def reconstruct_grid(strings, numbers, cell_data, template_data):
    """
    使用 Excel 模板的行列结构，将原始 strings/numbers 映射为二维数组。
    策略：
    1. 保留模板的第一行（表头）
    2. 对于数据行，尝试用原始数据填充
    3. 如果原始数据不足，保留模板中的值
    4. 优先使用 cell_data（field 6 的稀疏单元格数据）
    """
    if template_data is None:
        return None

    tmpl = template_data['data']
    nrows = len(tmpl)
    ncols = len(tmpl[0]) if tmpl else 0

    # 初始化结果数组（复制模板）
    result = []
    for r in range(nrows):
        result.append(list(tmpl[r]))

    # 优先使用 field 6 的稀疏单元格数据（最准确）
    # 但只有当 field 5.3 的数值明显不足时才使用（说明数据在 field 6 中）
    # field 6 的值可能是字符串索引（连续大数如 129,130...）或实际值（小离散数如 0,2,7...）
    # 判断标准：如果 field 5.3 的数字数量 < 预期单元格数的 20%，则使用 field 6
    expected_cells = (nrows - 1) * (ncols - 1)  # 除去表头行和标签列
    use_field6 = False
    if cell_data and len(numbers) < max(5, expected_cells * 0.2):
        # 检查 field 6 的值是否为字符串索引（连续大数）还是实际值
        vals = list(cell_data.values())
        if vals:
            is_string_index = all(v > 100 for v in vals) and max(vals) - min(vals) <= len(vals) + 10
            if not is_string_index:
                use_field6 = True

    if use_field6:
        for (row, col), val in cell_data.items():
            if row < nrows and col < ncols:
                result[row][col] = val
        return result

    if not strings or not numbers:
        return result

    # 尝试匹配：strings 中的行标签 vs 模板中的行标签
    # 模板第一行是表头，第一列是行标签
    # 分离 strings 中的行标签和列标题
    row_labels_from_data = []
    col_headers_from_data = []
    corner = strings[0] if strings else ''

    # 简单启发式：匹配模板中的行标签
    for i in range(1, len(strings)):
        s = strings[i]
        # 检查是否匹配模板中的行标签
        found = False
        for r in range(1, nrows):
            tmpl_label = str(tmpl[r][0]).strip() if tmpl[r][0] else ''
            if s == tmpl_label:
                row_labels_from_data.append((i, r, s))
                found = True
                break
        if not found:
            col_headers_from_data.append(s)

    # 按列分配数值
    # 先确定每列有多少个数值
    num_data_rows = len(row_labels_from_data)
    num_data_cols = len(col_headers_from_data) if col_headers_from_data else (ncols - 1)

    if num_data_rows == 0 or num_data_cols == 0:
        return result

    # 数值按列组织：先列0所有值，再列1所有值...
    total_cells = num_data_rows * num_data_cols
    if len(numbers) >= total_cells:
        # 有足够的数值，按列分配
        per_col = num_data_rows
        num_idx = 0
        for c in range(num_data_cols):
            for r_idx in range(min(per_col, num_data_rows)):
                if num_idx < len(numbers):
                    row_info = row_labels_from_data[r_idx]
                    result[row_info[1]][c + 1] = numbers[num_idx]
                    num_idx += 1
    else:
        # 数值不够，尝试按列非均匀分配
        # 每个列的数值数量可能不同（有些单元格为空）
        remaining = len(numbers)
        col_counts = []
        for c in range(num_data_cols):
            if c == num_data_cols - 1:
                col_counts.append(remaining)
            else:
                # 估计每列数量
                est = max(1, remaining // (num_data_cols - c))
                col_counts.append(est)
                remaining -= est

        num_idx = 0
        for c in range(num_data_cols):
            for r_idx in range(min(col_counts[c], num_data_rows)):
                if num_idx < len(numbers):
                    row_info = row_labels_from_data[r_idx]
                    if row_info[1] < nrows and c + 1 < ncols:
                        result[row_info[1]][c + 1] = numbers[num_idx]
                    num_idx += 1

    return result


def main():
    print("=" * 60)
    print("Tencent Docs -> JSON Sync Tool")
    print("=" * 60)
    print(f"Doc ID: {DOC_ID}")

    # 读取 Excel 模板
    templates = read_excel_template()

    all_data = {}
    success_count = 0

    for sheet_name, tab_id in TAB_IDS.items():
        print(f"\n[FETCH] {sheet_name} (tab={tab_id})")

        raw_data, error = fetch_raw_data(DOC_ID, tab_id)

        if error:
            print(f"  [FAIL] {error}")
            # 回退到 Excel 模板数据
            if templates and sheet_name in templates:
                all_data[sheet_name] = {
                    'rows': templates[sheet_name]['rows'],
                    'cols': templates[sheet_name]['cols'],
                    'data': templates[sheet_name]['data']
                }
                print(f"  [FALLBACK] Using Excel template data")
            continue

        strings = raw_data['strings']
        numbers = raw_data['numbers']
        cell_data = raw_data.get('cell_data', {})
        print(f"  Strings: {len(strings)}, Numbers: {len(numbers)}, Cells: {len(cell_data)}")

        # 使用模板重建网格
        template = templates.get(sheet_name) if templates else None
        grid = reconstruct_grid(strings, numbers, cell_data, template)

        if grid:
            all_data[sheet_name] = {
                'rows': len(grid),
                'cols': len(grid[0]) if grid else 0,
                'data': grid
            }
            print(f"  [OK] Grid: {len(grid)}r x {len(grid[0]) if grid else 0}c")
            success_count += 1
        else:
            # 回退到模板
            if template:
                all_data[sheet_name] = template
                print(f"  [FALLBACK] Cannot reconstruct grid, using template")
            else:
                all_data[sheet_name] = {
                    'rows': 0, 'cols': 0, 'data': [],
                    'error': 'Cannot reconstruct'
                }
                print(f"  [FAIL] Cannot reconstruct grid")

        time.sleep(0.3)

    # 保存
    output = {
        'source': 'tencent_docs',
        'doc_id': DOC_ID,
        'generated_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'sheets': all_data,
        'success_count': success_count,
        'total_sheets': len(TAB_IDS),
    }

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"[DONE] {success_count}/{len(TAB_IDS)} sheets synced")
    print(f"[FILE] {OUTPUT_FILE} ({os.path.getsize(OUTPUT_FILE):,} bytes)")


if __name__ == '__main__':
    main()